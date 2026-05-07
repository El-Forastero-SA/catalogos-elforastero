#!/usr/bin/env python3
"""
generar_desde_lista.py — Genera catalogo HTML desde una lista de precios Excel.

Lee un archivo Excel de lista de precios de El Forastero, obtiene imagenes
de productos desde Magento 2 API, y genera un catalogo HTML (+ PDF opcional)
con el mismo formato visual que generar_catalogo.py.

Uso:
    python3 generar_desde_lista.py --excel "/path/to/Lista_Mascotas.xlsx" --pdf

Requiere archivo .env en el mismo directorio con:
    MAGENTO_BASE_URL=https://tu-tienda.com
    MAGENTO_ACCESS_TOKEN=tu_token
"""

import argparse
import base64
import hashlib
import io
import json
import math
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from html import escape


# =============================================================================
# .env parser (sin dependencias externas)
# =============================================================================

def load_env(env_path):
    """Lee un archivo .env y retorna un dict con las variables."""
    env_vars = {}
    if not os.path.exists(env_path):
        return env_vars
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip()
            if len(value) >= 2 and value[0] == value[-1] and value[0] in ('"', "'"):
                value = value[1:-1]
            env_vars[key] = value
    return env_vars


# =============================================================================
# Magento API client
# =============================================================================

def magento_get_product(base_url, token, sku):
    """Obtiene un producto individual por SKU. Retorna dict o None si no existe."""
    encoded_sku = urllib.parse.quote(str(sku), safe="")
    url = f"{base_url.rstrip('/')}/rest/V1/products/{encoded_sku}"
    req = urllib.request.Request(url, method="GET")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Accept", "application/json")

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read().decode("utf-8")
            return json.loads(data)
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        print(f"  Error API Magento (HTTP {e.code}) para SKU {sku}: {e.reason}",
              file=sys.stderr)
        if body:
            print(f"    Detalle: {body[:300]}", file=sys.stderr)
        return None
    except urllib.error.URLError as e:
        print(f"  Error de conexion para SKU {sku}: {e.reason}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  Error inesperado para SKU {sku}: {e}", file=sys.stderr)
        return None


def fetch_product_image(base_url, token, sku):
    """Obtiene la URL de imagen de un producto por SKU. Retorna URL o None.

    Busca en orden: image > small_image > thumbnail (fallback).
    """
    product = magento_get_product(base_url, token, sku)
    if not product:
        return None

    # Extraer todos los atributos de imagen
    image_attrs = {}
    for attr in product.get("custom_attributes", []):
        code = attr.get("attribute_code")
        if code in ("image", "small_image", "thumbnail"):
            image_attrs[code] = attr.get("value")

    # Buscar en orden de prioridad
    for attr_name in ("image", "small_image", "thumbnail"):
        image_path = image_attrs.get(attr_name)
        if image_path and image_path != "no_selection":
            return f"{base_url.rstrip('/')}/pub/media/catalog/product{image_path}"

    return None


# =============================================================================
# Image auto-crop (remove white borders + transparent background)
# =============================================================================

# Umbral: un pixel se considera "blanco" si sus 3 canales RGB estan por encima.
# 240 tolera artifacts de compresion JPG sin comerse detalles reales del producto.
WHITE_THRESHOLD = 240

# Margen en pixeles que queda alrededor del producto despues de cropear.
# Chico para evitar el "doble contenedor" pero no al ras para no cortar detalles.
CROP_PADDING = 8

# Tamano maximo (lado mayor) despues de cropear. El contenedor del catalogo es
# 200x280px, asi que 600px = ~3x retina, suficiente para PDF crisp y peso bajo.
MAX_DIMENSION = 600

# Calidad JPG de salida. 85 es transparente a la vista para contenido fotografico
# sobre fondo blanco y reduce peso ~10x vs PNG.
JPG_QUALITY = 85


def _image_cache_dir(script_dir):
    cache_dir = os.path.join(script_dir, ".image_cache")
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir


def _cache_key(url):
    return hashlib.sha1(url.encode("utf-8")).hexdigest()


def download_image_bytes(url, timeout=30):
    """Descarga los bytes de una imagen. Retorna bytes o None."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "ForasteroCatalog/1.0"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except Exception as e:
        print(f"  Error descargando {url}: {e}", file=sys.stderr)
        return None


def process_image_crop_to_jpg(image_bytes):
    """Detecta bordes blancos, crop al contenido y exporta como JPG sobre blanco.

    Input: bytes de imagen (JPG, PNG, etc.)
    Output: bytes JPG con fondo blanco, cropeado tight al producto, o None.

    Uso de ImageChops.difference para encontrar el bbox del contenido de forma
    vectorizada (mucho mas rapido que iterar pixel por pixel en Python).
    """
    try:
        from PIL import Image, ImageChops
    except ImportError:
        print("  Error: Pillow no esta instalado. pip3 install Pillow", file=sys.stderr)
        return None

    try:
        src = Image.open(io.BytesIO(image_bytes))
        # Convertir a RGBA para preservar transparencia, despues componer sobre
        # blanco para obtener RGB. Hacer .convert('RGB') directo sobre un PNG
        # transparente deja los pixeles transparentes en NEGRO (default Pillow).
        if src.mode != "RGBA":
            src = src.convert("RGBA")
        w, h = src.size
        white_bg = Image.new("RGB", (w, h), (255, 255, 255))
        white_bg.paste(src, mask=src.split()[3])  # usa canal alpha como mascara
        img = white_bg

        # Crear imagen blanca de referencia y restar para obtener diferencia.
        # Pixeles casi blancos -> valores muy bajos -> los mandamos a 0 con el offset.
        bg = Image.new("RGB", (w, h), (255, 255, 255))
        diff = ImageChops.difference(img, bg)
        # Aplicar threshold: pixeles con diferencia menor a (255 - WHITE_THRESHOLD)
        # cuentan como blanco y no entran al bbox.
        offset = 255 - WHITE_THRESHOLD
        diff = ImageChops.add(diff, diff, scale=1.0, offset=-offset)
        bbox = diff.getbbox()
        if bbox is None:
            return None

        # Expandir bbox con padding, respetando limites
        left, top, right, bottom = bbox
        left = max(0, left - CROP_PADDING)
        top = max(0, top - CROP_PADDING)
        right = min(w, right + CROP_PADDING)
        bottom = min(h, bottom + CROP_PADDING)

        cropped = img.crop((left, top, right, bottom))

        # Redimensionar si excede MAX_DIMENSION (mantiene aspect ratio)
        cw, ch = cropped.size
        if max(cw, ch) > MAX_DIMENSION:
            scale = MAX_DIMENSION / max(cw, ch)
            new_size = (int(cw * scale), int(ch * scale))
            cropped = cropped.resize(new_size, Image.LANCZOS)

        out = io.BytesIO()
        cropped.save(out, format="JPEG", quality=JPG_QUALITY, optimize=True,
                     progressive=True)
        return out.getvalue()
    except Exception as e:
        print(f"  Error procesando imagen: {e}", file=sys.stderr)
        return None


def get_processed_image_data_uri(url, script_dir):
    """Descarga, cachea y procesa una imagen. Retorna data URI base64 o None.

    Usa cache en disco (.image_cache/) para no reprocesar entre runs.
    """
    cache_dir = _image_cache_dir(script_dir)
    cache_file = os.path.join(cache_dir, _cache_key(url) + ".jpg")

    if os.path.exists(cache_file):
        with open(cache_file, "rb") as f:
            jpg_bytes = f.read()
    else:
        raw = download_image_bytes(url)
        if raw is None:
            return None
        jpg_bytes = process_image_crop_to_jpg(raw)
        if jpg_bytes is None:
            return None
        with open(cache_file, "wb") as f:
            f.write(jpg_bytes)

    b64 = base64.b64encode(jpg_bytes).decode("ascii")
    return f"data:image/jpeg;base64,{b64}"


# =============================================================================
# Excel reader
# =============================================================================

def get_hex_color(cell):
    """Extrae el color hex del fondo de una celda de openpyxl."""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            color = fill.fgColor
            # Color RGB directo
            if color.rgb and color.rgb != "00000000":
                rgb = str(color.rgb)
                if len(rgb) == 8:  # AARRGGBB format
                    return "#" + rgb[2:]
                if len(rgb) == 6:
                    return "#" + rgb
            # Color con theme (intentar leer el tint/index)
            if color.theme is not None:
                # openpyxl theme colors — mapeo basico de los mas comunes
                # No siempre es posible resolver sin el theme XML completo
                pass
    except Exception:
        pass
    return None


def clean_category_name(raw_name):
    """Limpia el nombre de categoria del Excel para mostrar en el catalogo.

    Ejemplos:
        "MASCOTAS - ALIMENTOS SUPER PREMIUM" -> "Alimentos Super Premium"
        "CONSUMO - BEBIDAS" -> "Bebidas"
        "Cafe" -> "Cafe"
    """
    if not raw_name:
        return "Productos"

    name = str(raw_name).strip()

    # Si tiene " - ", tomar la parte despues del guion (que es la sub-categoria)
    if " - " in name:
        parts = name.split(" - ", 1)
        name = parts[1].strip()

    # Convertir a title case
    name = name.title()

    return name


def read_excel_products(excel_path):
    """Lee productos desde un Excel de lista de precios.

    Retorna: (category_name, header_color, products_list)
    donde products_list es [{sku, name}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl no esta instalado.", file=sys.stderr)
        print("  Instalar con: pip3 install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(excel_path):
        print(f"Error: no se encontro el archivo: {excel_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    # Leer nombre de categoria de fila 2 (puede estar en A2 mergeado o B2)
    raw_category = ws.cell(row=2, column=1).value or ws.cell(row=2, column=2).value
    category_name = clean_category_name(raw_category)

    # Leer color del header de A1
    header_color = get_hex_color(ws.cell(row=1, column=1))

    # Iterar filas desde la 6 (despues de los headers)
    products = []
    for row in ws.iter_rows(min_row=6, max_col=2):
        cell_a = row[0]  # Columna A: SKU
        cell_b = row[1]  # Columna B: Descripcion

        # Verificar que columna A tiene un valor (SKU)
        val_a = cell_a.value
        if val_a is None:
            continue

        # Verificar que es numerico (SKU)
        val_a_str = str(val_a).strip()
        if not val_a_str:
            continue

        # Intentar interpretar como numero (algunos vienen como int, otros como str)
        try:
            # Quitar decimales si vino como float (ej: 7401.0 -> "7401")
            numeric_val = int(float(val_a_str))
            sku = str(numeric_val).zfill(5)
        except (ValueError, TypeError):
            # No es numerico — es una fila separadora o nota
            continue

        # Leer descripcion
        val_b = cell_b.value
        if not val_b:
            continue

        name = str(val_b).strip()
        if not name:
            continue

        products.append({
            "sku": sku,
            "name": name,
        })

    wb.close()
    return category_name, header_color, products


# =============================================================================
# Utilidades de texto
# =============================================================================

def split_product_name(name):
    """
    Divide el nombre del producto en dos lineas:
    Linea 1: marca + producto principal
    Linea 2: variante/presentacion

    Heuristica: busca el ultimo " x " o " - " para separar.
    """
    idx_dash = name.rfind(" - ")
    idx_x = name.lower().rfind(" x ")
    split_idx = max(idx_dash, idx_x)

    if split_idx > 0 and split_idx < len(name) - 3:
        line1 = name[:split_idx].strip()
        line2 = name[split_idx + 3:].strip()
        return line1, line2

    return name, ""


# =============================================================================
# SVG Icons (inline, color blanco para fondos oscuros)
# =============================================================================

SVG_ICON_WEB = (
    '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" viewBox="0 0 24 24" fill="white">'
    '<path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2zm-1 17.93'
    'c-3.95-.49-7-3.85-7-7.93 0-.62.08-1.21.21-1.79L9 15v1c0 1.1.9 2 2 2v1.93zm6.9-2.54'
    'c-.26-.81-1-1.39-1.9-1.39h-1v-3c0-.55-.45-1-1-1H8v-2h2c.55 0 1-.45 1-1V7h2c1.1 0 '
    '2-.9 2-2v-.41c2.93 1.19 5 4.06 5 7.41 0 2.08-.8 3.97-2.1 5.39z"/>'
    '</svg>'
)

SVG_ICON_INSTAGRAM = (
    '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" viewBox="0 0 24 24" fill="white">'
    '<path d="M7.8 2h8.4C19.4 2 22 4.6 22 7.8v8.4a5.8 5.8 0 0 1-5.8 5.8H7.8C4.6 22 2 '
    '19.4 2 16.2V7.8A5.8 5.8 0 0 1 7.8 2m-.2 2A3.6 3.6 0 0 0 4 7.6v8.8C4 18.39 5.61 20 '
    '7.6 20h8.8a3.6 3.6 0 0 0 3.6-3.6V7.6C20 5.61 18.39 4 16.4 4H7.6m9.65 1.5a1.25 1.25 '
    '0 0 1 1.25 1.25A1.25 1.25 0 0 1 17.25 8 1.25 1.25 0 0 1 16 6.75a1.25 1.25 0 0 1 '
    '1.25-1.25M12 7a5 5 0 0 1 5 5 5 5 0 0 1-5 5 5 5 0 0 1-5-5 5 5 0 0 1 5-5m0 2a3 3 0 '
    '0 0-3 3 3 3 0 0 0 3 3 3 3 0 0 0 3-3 3 3 0 0 0-3-3z"/>'
    '</svg>'
)

SVG_ICON_WHATSAPP = (
    '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" viewBox="0 0 24 24" fill="white">'
    '<path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15'
    '-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475'
    '-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52'
    '.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207'
    '-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297'
    '-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487'
    '.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413'
    '.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 '
    '0 0 1-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 0 1-1.51-5.26'
    'c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 0 1 '
    '2.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0 0 12.05 '
    '0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 '
    '11.882 0 0 0 5.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 0 0'
    '-3.48-8.413z"/>'
    '</svg>'
)

SVG_ICON_WHATSAPP_GREEN = SVG_ICON_WHATSAPP.replace('fill="white"', 'fill="#25D366"')


# =============================================================================
# Logo blanco (base64 PNG)
# =============================================================================

def _load_b64_file(filename):
    """Carga un archivo base64 desde el directorio del script."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    if os.path.exists(path):
        with open(path, "r") as f:
            return f.read().strip()
    return None

LOGO_WHITE_B64 = _load_b64_file("logo_white_b64.txt")
LOGO_BLACK_B64 = _load_b64_file("logo_black_b64.txt")


# =============================================================================
# HTML Generation (mismo formato que generar_catalogo.py)
# =============================================================================

SVG_BG_PATTERN = (
    "data:image/svg+xml,%3Csvg width='160' height='160' xmlns='http://www.w3.org/2000/svg'"
    "%3E%3Cdefs%3E%3Cstyle%3E.t%7Bopacity:.055;fill:none;stroke-width:1.5%7D"
    ".d%7Bopacity:.04%7D%3C/style%3E%3C/defs%3E%3Cpolygon class='t' "
    "points='20,8 28,22 12,22' stroke='%23555'/%3E%3Cpolygon class='t' "
    "points='75,15 83,29 67,29' stroke='%23999' transform='rotate(20 75 22)'/%3E"
    "%3Cpolygon class='t' points='130,10 138,24 122,24' stroke='%23555' "
    "transform='rotate(-15 130 17)'/%3E%3Cpolygon class='t' points='45,55 53,69 37,69' "
    "stroke='%23777' transform='rotate(35 45 62)'/%3E%3Cpolygon class='t' "
    "points='110,50 118,64 102,64' stroke='%23555' transform='rotate(-25 110 57)'/%3E"
    "%3Cpolygon class='t' points='15,95 23,109 7,109' stroke='%23999' "
    "transform='rotate(10 15 102)'/%3E%3Cpolygon class='t' points='85,90 93,104 77,104' "
    "stroke='%23555' transform='rotate(-40 85 97)'/%3E%3Cpolygon class='t' "
    "points='140,85 148,99 132,99' stroke='%23777'/%3E%3Cpolygon class='t' "
    "points='50,130 58,144 42,144' stroke='%23555' transform='rotate(30 50 137)'/%3E"
    "%3Cpolygon class='t' points='115,135 123,149 107,149' stroke='%23999' "
    "transform='rotate(-10 115 142)'/%3E%3Ccircle class='d' cx='35' cy='35' r='2' "
    "fill='%23888'/%3E%3Ccircle class='d' cx='100' cy='75' r='1.8' fill='%23888'/%3E"
    "%3Ccircle class='d' cx='60' cy='110' r='2' fill='%23888'/%3E%3Crect class='d' "
    "x='145' y='40' width='4' height='4' fill='%23888' transform='rotate(45 147 42)'/%3E"
    "%3Crect class='d' x='25' y='75' width='3' height='3' fill='%23888' "
    "transform='rotate(30 26.5 76.5)'/%3E%3C/svg%3E"
)


FONTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
FONTS_SPEC = [
    ("Montserrat",       "400", "Montserrat-400.woff2"),
    ("Montserrat",       "500", "Montserrat-500.woff2"),
    ("Montserrat",       "600", "Montserrat-600.woff2"),
    ("Montserrat",       "700", "Montserrat-700.woff2"),
    ("Montserrat",       "800", "Montserrat-800.woff2"),
    ("Playfair Display", "700", "PlayfairDisplay-700.woff2"),
    ("Playfair Display", "800", "PlayfairDisplay-800.woff2"),
]


def render_fonts_css():
    """Embebe los .woff2 como data:URI base64 en @font-face — evita problemas
    de carga via file:// en Chromium/Linux que terminan en fallback Type 3."""
    blocks = []
    for family, weight, fname in FONTS_SPEC:
        path = os.path.join(FONTS_DIR, fname)
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        blocks.append(
            f"@font-face {{ font-family: '{family}'; font-weight: {weight}; "
            f"font-style: normal; "
            f"src: url('data:font/woff2;base64,{b64}') format('woff2'); }}"
        )
    return "\n    ".join(blocks)


def generate_css(category_color):
    """Genera el CSS completo embebido, basado en el template."""
    return f"""
    :root {{
      --category-color: {category_color};
      --bg-light: #E8E5E0;
      --text-dark: #333333;
      --text-mid: #555555;
      --text-light: #777777;
      --white: #FFFFFF;
      --page-w: 1080px;
      --page-h: 1920px;
    }}

    *, *::before, *::after {{
      margin: 0;
      padding: 0;
      box-sizing: border-box;
    }}

    html {{
      -webkit-print-color-adjust: exact;
      print-color-adjust: exact;
    }}

    body {{
      font-family: 'Montserrat', Arial, sans-serif;
      background: white;
      color: var(--text-dark);
      margin: 0;
      padding: 0;
    }}

    @page {{
      size: 1080px 1920px;
      margin: 0;
    }}

    .page {{
      position: relative;
      width: var(--page-w);
      height: var(--page-h);
      background: var(--white);
      overflow: hidden;
    }}

    @media screen {{
      body {{ background: #ccc; }}
      .page {{ margin: 20px auto; box-shadow: 0 2px 12px rgba(0,0,0,0.15); }}
    }}

    .content-area {{
      position: relative;
      background-color: var(--bg-light);
      background-image: url("{SVG_BG_PATTERN}");
      background-repeat: repeat;
      width: 100%;
      height: 1516px;
      padding: 36px 40px 24px;
    }}

    /* ---- Header ---- */
    .page-header {{
      position: relative;
      width: 100%;
      height: 154px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      background: var(--white);
      z-index: 10;
    }}

    .header-logo {{
      display: flex;
      align-items: center;
      gap: 14px;
      padding-left: 40px;
      flex-shrink: 0;
    }}

    .header-logo-img {{
      height: 90px;
      width: auto;
    }}

    .header-bar {{
      position: relative;
      height: 100%;
      flex: 1;
      max-width: 620px;
      display: flex;
      align-items: center;
      justify-content: center;
      background: var(--category-color);
      clip-path: polygon(80px 0, 100% 0, 100% 100%, 0 100%);
      padding-left: 100px;
      padding-right: 40px;
    }}

    .header-bar-title {{
      color: var(--white);
      font-weight: 800;
      font-size: 28px;
      letter-spacing: 3px;
      text-transform: uppercase;
      white-space: nowrap;
    }}

    /* ---- Product Grid ---- */
    .product-grid {{
      display: grid;
      grid-template-columns: repeat(4, 1fr);
      grid-template-rows: repeat(3, 1fr);
      gap: 20px 16px;
      width: 100%;
      height: 100%;
    }}

    .product-card {{
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: flex-start;
      padding: 12px 8px 8px;
      page-break-inside: avoid;
    }}

    .product-image-wrap {{
      width: 200px;
      height: 280px;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 12px;
      background: #ffffff;
      border-radius: 12px;
      flex-shrink: 0;
    }}

    .product-image {{
      max-width: 100%;
      max-height: 100%;
      width: auto;
      height: auto;
      object-fit: contain;
    }}

    .product-image-placeholder {{
      width: 80%;
      height: 85%;
      background: linear-gradient(135deg, #d5d0ca 0%, #c2bcb4 100%);
      border-radius: 12px;
      display: flex;
      align-items: center;
      justify-content: center;
      color: #a09a92;
      font-size: 14px;
      font-weight: 600;
    }}

    .product-info {{
      width: 100%;
      text-align: center;
      padding: 4px 2px 0;
      line-height: 1.3;
    }}

    .product-name {{
      font-size: 13px;
      font-weight: 700;
      color: var(--text-dark);
      margin-bottom: 1px;
      display: -webkit-box;
      -webkit-line-clamp: 3;
      -webkit-box-orient: vertical;
      overflow: hidden;
      line-height: 1.25;
      word-break: break-word;
    }}

    .product-variant {{
      font-size: 12px;
      font-weight: 600;
      color: var(--text-mid);
      margin-bottom: 1px;
      line-height: 1.2;
    }}

    .product-meta {{
      font-size: 12px;
      font-weight: 500;
      color: var(--text-light);
    }}

    /* ---- Footer ---- */
    .page-footer {{
      width: 100%;
      height: 250px;
      background: var(--category-color);
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      gap: 10px;
      padding: 16px 40px;
    }}

    .footer-logo-img {{
      height: 40px;
      width: auto;
    }}

    .footer-contact-row {{
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 28px;
      color: var(--white);
      font-size: 16px;
      font-weight: 500;
    }}

    .footer-contact-row svg {{
      width: 20px;
      height: 20px;
      fill: var(--white);
      flex-shrink: 0;
      vertical-align: middle;
      margin-right: 6px;
    }}

    .footer-contact-item {{
      display: flex;
      align-items: center;
    }}

    .footer-phones {{
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 24px;
      color: var(--white);
      font-size: 14px;
      font-weight: 500;
      flex-wrap: wrap;
    }}

    .footer-phone-item {{
      display: flex;
      align-items: center;
      gap: 6px;
    }}

    .footer-phone-item svg {{
      width: 18px;
      height: 18px;
      fill: var(--white);
      flex-shrink: 0;
    }}

    .footer-phone-branch {{
      font-weight: 700;
    }}

    .footer-divider {{
      width: 60%;
      height: 1px;
      background: rgba(255,255,255,0.25);
    }}

    /* ---- Cover ---- */
    .page-cover {{
      background: var(--category-color);
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      text-align: center;
    }}

    .cover-logo {{
      width: 520px;
      height: auto;
      margin-bottom: 60px;
      filter: drop-shadow(0 4px 20px rgba(0,0,0,0.15));
    }}

    .cover-divider {{
      width: 160px;
      height: 2px;
      background: rgba(255,255,255,0.35);
      margin-bottom: 60px;
    }}

    .cover-category {{
      font-size: 72px;
      font-weight: 800;
      color: var(--white);
      text-transform: uppercase;
      letter-spacing: 6px;
      line-height: 1.1;
      margin-bottom: 16px;
      padding: 0 60px;
    }}

    .cover-subcategory {{
      font-size: 28px;
      font-weight: 600;
      color: rgba(255,255,255,0.75);
      letter-spacing: 3px;
      text-transform: uppercase;
      margin-bottom: 120px;
    }}

    .cover-contact {{
      display: flex;
      align-items: center;
      gap: 28px;
      color: rgba(255,255,255,0.7);
      font-size: 16px;
      font-weight: 500;
    }}

    .cover-contact svg {{
      width: 20px;
      height: 20px;
      fill: rgba(255,255,255,0.7);
      margin-right: 6px;
      vertical-align: middle;
    }}

    /* ---- Back Page ---- */
    .page-back {{
      background: var(--white);
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      text-align: center;
    }}

    .back-logo-img {{
      width: 400px;
      height: auto;
      margin-bottom: 60px;
    }}

    .back-divider {{
      width: 200px;
      height: 2px;
      background: #ddd;
      margin-bottom: 60px;
    }}

    .back-branches {{
      display: flex;
      gap: 60px;
      margin-bottom: 60px;
    }}

    .back-branch {{
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: 10px;
    }}

    .back-branch-name {{
      font-size: 20px;
      font-weight: 700;
      color: var(--text-dark);
      text-transform: uppercase;
      letter-spacing: 2px;
    }}

    .back-branch-phone {{
      display: flex;
      align-items: center;
      gap: 8px;
      font-size: 18px;
      font-weight: 500;
      color: var(--text-mid);
    }}

    .back-branch-phone svg {{
      width: 22px;
      height: 22px;
      fill: #25D366;
    }}

    .back-tagline {{
      font-size: 26px;
      font-weight: 700;
      color: var(--category-color);
      margin-top: 20px;
    }}

    .back-web {{
      display: flex;
      align-items: center;
      gap: 24px;
      margin-top: 40px;
      font-size: 18px;
      color: var(--text-light);
      font-weight: 500;
    }}

    .back-web svg {{
      width: 20px;
      height: 20px;
      fill: var(--text-light);
      margin-right: 6px;
      vertical-align: middle;
    }}
    """


def render_product_card(product):
    """Genera el HTML de una tarjeta de producto individual."""
    name_line1, name_line2 = split_product_name(product["name"])

    if product.get("image_url"):
        image_html = (
            f'<img class="product-image" '
            f'src="{escape(product["image_url"], quote=True)}" '
            f'alt="{escape(name_line1)}">'
        )
    else:
        image_html = '<div class="product-image-placeholder">FOTO</div>'

    variant_html = ""
    if name_line2:
        variant_html = f'<div class="product-variant">{escape(name_line2)}</div>'

    return f"""
        <div class="product-card">
          <div class="product-image-wrap">
            {image_html}
          </div>
          <div class="product-info">
            <div class="product-name">{escape(name_line1)}</div>
            {variant_html}
            <div class="product-meta">C&oacute;d. {escape(product["sku"])}</div>
          </div>
        </div>"""


def render_cover_page(category_name):
    """Genera la pagina de portada."""
    if LOGO_WHITE_B64:
        logo_html = (
            f'<img class="cover-logo" '
            f'src="data:image/png;base64,{LOGO_WHITE_B64}" '
            f'alt="El Forastero - Desde 1959">'
        )
    else:
        # Fallback texto si no se encuentra el logo
        logo_html = (
            '<div style="font-family:Playfair Display,serif;font-size:56px;'
            'font-weight:800;color:white;letter-spacing:6px;font-variant:small-caps;'
            'margin-bottom:60px;">El Forastero</div>'
        )

    return f"""
  <div class="page page-cover">
    {logo_html}
    <div class="cover-divider"></div>
    <div class="cover-category">{escape(category_name)}</div>
    <div class="cover-subcategory">Portfolio de Productos</div>
    <div class="cover-contact">
      <span>{SVG_ICON_WEB} www.elforastero.com.ar</span>
      <span>{SVG_ICON_INSTAGRAM} @elforasterosa</span>
    </div>
  </div>"""


def render_product_page(products_chunk):
    """Genera una pagina de productos (hasta 12 productos)."""
    cards_html = "\n".join(render_product_card(p) for p in products_chunk)

    return f"""
  <div class="page">
    <div class="page-header">
      <div class="header-logo">
        <img class="header-logo-img" src="data:image/png;base64,{LOGO_BLACK_B64}" alt="El Forastero">
      </div>
      <div class="header-bar">
        <span class="header-bar-title">Portfolio de Productos</span>
      </div>
    </div>

    <div class="content-area">
      <div class="product-grid">
{cards_html}
      </div>
    </div>

    <div class="page-footer">
      <img class="footer-logo-img" src="data:image/png;base64,{LOGO_WHITE_B64}" alt="El Forastero">
      <div class="footer-contact-row">
        <span class="footer-contact-item">
          {SVG_ICON_WEB} www.elforastero.com.ar
        </span>
        <span class="footer-contact-item">
          {SVG_ICON_INSTAGRAM} @elforasterosa
        </span>
      </div>
      <div class="footer-divider"></div>
      <div class="footer-phones">
        <div class="footer-phone-item">
          {SVG_ICON_WHATSAPP}
          <span>298 4222 800</span>
          <span class="footer-phone-branch">CERVANTES</span>
        </div>
        <div class="footer-phone-item">
          {SVG_ICON_WHATSAPP}
          <span>298 4222 800</span>
          <span class="footer-phone-branch">SANTA ROSA</span>
        </div>
        <div class="footer-phone-item">
          {SVG_ICON_WHATSAPP}
          <span>294 441-0174</span>
          <span class="footer-phone-branch">BARILOCHE</span>
        </div>
        <div class="footer-phone-item">
          {SVG_ICON_WHATSAPP}
          <span>2396 44-0420</span>
          <span class="footer-phone-branch">PEHUAJO</span>
        </div>
      </div>
    </div>
  </div>"""


def render_back_page():
    """Genera la contraportada."""
    return f"""
  <div class="page page-back">
    <img class="back-logo-img" src="data:image/png;base64,{LOGO_BLACK_B64}" alt="El Forastero - Desde 1959">
    <div class="back-divider"></div>

    <div class="back-branches">
      <div class="back-branch">
        <div class="back-branch-name">Cervantes</div>
        <div class="back-branch-phone">
          {SVG_ICON_WHATSAPP_GREEN}
          298 4222 800
        </div>
      </div>
      <div class="back-branch">
        <div class="back-branch-name">Santa Rosa</div>
        <div class="back-branch-phone">
          {SVG_ICON_WHATSAPP_GREEN}
          298 4222 800
        </div>
      </div>
      <div class="back-branch">
        <div class="back-branch-name">Bariloche</div>
        <div class="back-branch-phone">
          {SVG_ICON_WHATSAPP_GREEN}
          294 441-0174
        </div>
      </div>
      <div class="back-branch">
        <div class="back-branch-name">Pehuajo</div>
        <div class="back-branch-phone">
          {SVG_ICON_WHATSAPP_GREEN}
          2396 44-0420
        </div>
      </div>
    </div>

    <div class="back-tagline">&iexcl;Estamos listos para ayudarte a crecer!</div>

    <div class="back-web">
      <span>
        {SVG_ICON_WEB.replace('fill="white"', 'fill="#777777"')}
        www.elforastero.com.ar
      </span>
      <span>
        {SVG_ICON_INSTAGRAM.replace('fill="white"', 'fill="#777777"')}
        @elforasterosa
      </span>
    </div>
  </div>"""


def generate_html(category_name, category_color, products, products_per_page):
    """Genera el documento HTML completo del catalogo."""
    css = generate_css(category_color)

    num_pages = math.ceil(len(products) / products_per_page) if products else 0
    pages_html = []

    for i in range(num_pages):
        start = i * products_per_page
        end = start + products_per_page
        chunk = products[start:end]
        pages_html.append(render_product_page(chunk))

    cover = render_cover_page(category_name)
    back = render_back_page()
    all_pages = cover + "\n" + "\n".join(pages_html) + "\n" + back

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=1080">
  <title>El Forastero - {escape(category_name)} - Portfolio de Productos</title>
  <style>
{render_fonts_css()}
{css}
  </style>
</head>
<body>
{all_pages}
</body>
</html>"""

    return html, num_pages


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Genera catalogo HTML desde una lista de precios Excel."
    )
    parser.add_argument(
        "--excel",
        required=True,
        help="Ruta al archivo Excel de lista de precios",
    )
    parser.add_argument(
        "--category-name",
        default=None,
        help="Nombre para mostrar en el catalogo (default: leer de celda B2 del Excel)",
    )
    parser.add_argument(
        "--category-color",
        default=None,
        help="Color hex para la categoria (default: leer del header del Excel, o #BC1221)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Nombre del archivo de salida (default: catalogo_{nombre}.html)",
    )
    parser.add_argument(
        "--products-per-page",
        type=int,
        default=12,
        help="Productos por pagina (default: 12)",
    )
    parser.add_argument(
        "--pdf",
        action="store_true",
        help="Generar tambien PDF (requiere playwright instalado)",
    )

    args = parser.parse_args()

    # --- Cargar .env ---
    script_dir = os.path.dirname(os.path.abspath(__file__))
    env_path = os.path.join(script_dir, ".env")
    env_vars = load_env(env_path)

    base_url = env_vars.get("MAGENTO_BASE_URL", "").rstrip("/")
    token = env_vars.get("MAGENTO_ACCESS_TOKEN", "")

    if not base_url:
        print("Error: MAGENTO_BASE_URL no definida en .env", file=sys.stderr)
        sys.exit(1)
    if not token:
        print("Error: MAGENTO_ACCESS_TOKEN no definido en .env", file=sys.stderr)
        sys.exit(1)

    # --- Leer Excel ---
    excel_path = os.path.abspath(args.excel)
    print(f"Leyendo Excel: {excel_path}")

    excel_category, excel_color, products = read_excel_products(excel_path)

    # Determinar nombre de categoria
    category_name = args.category_name or excel_category
    print(f"Categoria: {category_name}")

    # Determinar color
    category_color = args.category_color or excel_color or "#BC1221"
    print(f"Color: {category_color}")

    print(f"Productos encontrados en Excel: {len(products)}")

    if not products:
        print("No se encontraron productos en el Excel. Abortando.", file=sys.stderr)
        sys.exit(1)

    # --- Obtener imagenes de Magento ---
    print(f"\nObteniendo imagenes de Magento API...")
    image_cache = {}  # sku -> image_url o None
    total = len(products)
    not_found = 0

    for idx, product in enumerate(products, 1):
        sku = product["sku"]

        # Usar cache si ya se busco este SKU
        if sku in image_cache:
            product["image_url"] = image_cache[sku]
        else:
            # Mostrar progreso
            sys.stdout.write(f"\r  Obteniendo imagenes: {idx}/{total}...")
            sys.stdout.flush()

            image_url = fetch_product_image(base_url, token, sku)
            image_cache[sku] = image_url
            product["image_url"] = image_url

            if image_url is None:
                not_found += 1

    print(f"\r  Obteniendo imagenes: {total}/{total} - listo.")
    if not_found:
        print(f"  Productos sin imagen: {not_found} (se excluyen del catalogo)")

    # --- Procesar imagenes: crop + transparencia (elimina doble contenedor) ---
    print(f"\nProcesando imagenes (crop + transparencia)...")
    products_with_images = [p for p in products if p.get("image_url")]
    proc_total = len(products_with_images)
    proc_failed = 0
    for idx, product in enumerate(products_with_images, 1):
        sys.stdout.write(f"\r  Procesando: {idx}/{proc_total}...")
        sys.stdout.flush()
        data_uri = get_processed_image_data_uri(product["image_url"], script_dir)
        if data_uri:
            product["image_url"] = data_uri
        else:
            product["image_url"] = None
            proc_failed += 1
    print(f"\r  Procesando: {proc_total}/{proc_total} - listo.")
    if proc_failed:
        print(f"  Fallos de procesamiento: {proc_failed} (se excluyen del catalogo)")

    # --- Filtrar productos sin imagen ---
    products = [p for p in products if p.get("image_url")]
    print(f"  Productos con imagen: {len(products)}")

    if not products:
        print("Ningun producto tiene imagen. Abortando.", file=sys.stderr)
        sys.exit(1)

    # --- Generar HTML ---
    html_content, num_pages = generate_html(
        category_name,
        category_color,
        products,
        args.products_per_page,
    )

    # --- Guardar archivo ---
    if args.output:
        output_path = args.output
    else:
        safe_name = category_name.lower().replace(" ", "_")
        safe_name = "".join(c for c in safe_name if c.isalnum() or c == "_")
        output_path = f"catalogo_{safe_name}.html"

    if not os.path.isabs(output_path):
        output_path = os.path.join(script_dir, output_path)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    # --- Generar PDF si se pidio ---
    pdf_path = None
    if args.pdf:
        pdf_path = output_path.replace(".html", ".pdf")
        print(f"Generando PDF...")
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page()
                page.goto(f"file://{output_path}", wait_until="networkidle",
                          timeout=120000)
                page.wait_for_timeout(3000)
                page.pdf(
                    path=pdf_path,
                    width="1080px",
                    height="1920px",
                    margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
                    print_background=True,
                    prefer_css_page_size=True,
                )
                browser.close()
            print(f"PDF generado: {pdf_path}")
        except ImportError:
            print("Error: playwright no esta instalado.", file=sys.stderr)
            print("  Instalar con: pip3 install playwright && "
                  "python3 -m playwright install chromium", file=sys.stderr)
            pdf_path = None
        except Exception as e:
            print(f"Error generando PDF: {e}", file=sys.stderr)
            pdf_path = None

    # --- Resumen ---
    total_pages = 1 + num_pages + 1  # portada + producto pages + contraportada
    print("")
    print("=" * 50)
    print(f"  Catalogo generado exitosamente")
    print(f"  Categoria:    {category_name}")
    print(f"  Color:        {category_color}")
    print(f"  Productos:    {len(products)}")
    print(f"  Sin imagen:   {not_found}")
    print(f"  Paginas:      {total_pages} (portada + {num_pages} producto + contraportada)")
    print(f"  HTML:         {output_path}")
    if pdf_path:
        print(f"  PDF:          {pdf_path}")
    print("=" * 50)


if __name__ == "__main__":
    main()
